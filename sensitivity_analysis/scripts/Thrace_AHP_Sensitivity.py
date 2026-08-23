# Cluster-specific AHP sensitivity analysis for the Northern Thrace risk model.
# Tests one adjacent Saaty level in both directions for each pairwise comparison.

import csv
import math
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


# ============================================================================
# FILE PATHS AND VARIABLES
# ============================================================================

Script_folder = Path(__file__).resolve().parent
Analysis_folder = Script_folder.parent
Repository_folder = Analysis_folder.parent

AHP_files = {
    1: Repository_folder / "AHP" / "matrices" / "Thrace_AHP_C1.xlsx",
    2: Repository_folder / "AHP" / "matrices" / "Thrace_AHP_C2.xlsx",
    3: Repository_folder / "AHP" / "matrices" / "Thrace_AHP_C3.xlsx",
}
Site_file = Analysis_folder / "inputs" / "Sites_sensitivity_input.csv"
Output_folder = Analysis_folder / "results"

Criteria = [
    "Seismic", "Wildfire", "Flood", "Soil Erosion",
    "Landslide", "Asset Vulnerability",
]
Score_fields = {
    "Seismic": "SEIS_SCR",
    "Wildfire": "FWI_SCR",
    "Flood": "FLOOD_SCR",
    "Soil Erosion": "RUSLE_SCR",
    "Landslide": "ELSUS_SCR",
    "Asset Vulnerability": "ASSET_SCR",
}
Saaty_scale = [
    1 / 9, 1 / 8, 1 / 7, 1 / 6, 1 / 5, 1 / 4, 1 / 3, 1 / 2,
    1, 2, 3, 4, 5, 6, 7, 8, 9,
]
Risk_classes = ["Very Low", "Low", "Moderate", "High", "Very High"]
CR_threshold = 0.10
RI = 1.24


# ============================================================================
# GENERAL FUNCTIONS
# ============================================================================

def convert_to_number(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def format_saaty_value(value):
    for denominator in range(2, 10):
        if math.isclose(value, 1 / denominator, abs_tol=1e-9):
            return f"1/{denominator}"
    if math.isclose(value, round(value), abs_tol=1e-9):
        return str(int(round(value)))
    return f"{value:.6g}"


def classify_risk(score):
    if score < 2:
        return "Very Low"
    if score < 4:
        return "Low"
    if score < 6:
        return "Moderate"
    if score < 8:
        return "High"
    return "Very High"


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8-sig") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


# ============================================================================
# AHP MATRIX AND WEIGHT CALCULATIONS
# ============================================================================

def read_ahp_matrix(ahp_file):
    workbook = load_workbook(ahp_file, data_only=True, read_only=True)
    sheet = workbook.active
    header_row = start_column = None

    for row_number in range(1, min(sheet.max_row, 25) + 1):
        for column_number in range(1, sheet.max_column - len(Criteria) + 2):
            header_values = [
                sheet.cell(row_number, column_number + criterion_index).value
                for criterion_index in range(len(Criteria))
            ]
            if header_values == Criteria:
                header_row, start_column = row_number, column_number
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(f"Could not find the AHP criterion headers in {ahp_file.name}.")

    criterion_row_labels = [
        sheet.cell(header_row + 1 + criterion_index, start_column - 1).value
        for criterion_index in range(len(Criteria))
    ]
    if criterion_row_labels != Criteria:
        raise ValueError(f"Criterion rows do not match the expected order in {ahp_file.name}.")

    pairwise_matrix = []
    for row_index in range(len(Criteria)):
        matrix_row = [
            convert_to_number(sheet.cell(header_row + 1 + row_index, start_column + column_index).value)
            for column_index in range(len(Criteria))
        ]
        if None in matrix_row:
            raise ValueError(f"Blank matrix value found in {ahp_file.name}.")
        pairwise_matrix.append(matrix_row)

    for row_index in range(len(Criteria)):
        if not math.isclose(pairwise_matrix[row_index][row_index], 1.0, abs_tol=1e-8):
            raise ValueError(f"The diagonal of {ahp_file.name} contains a value other than 1.")
        for column_index in range(row_index + 1, len(Criteria)):
            reciprocal_value = pairwise_matrix[row_index][column_index] * pairwise_matrix[column_index][row_index]
            if not math.isclose(reciprocal_value, 1.0, abs_tol=1e-7):
                raise ValueError(f"The AHP matrix in {ahp_file.name} is not reciprocal.")
    return pairwise_matrix


def calculate_ahp(pairwise_matrix):
    criterion_count = len(pairwise_matrix)
    column_sums = [
        sum(pairwise_matrix[row_index][column_index] for row_index in range(criterion_count))
        for column_index in range(criterion_count)
    ]
    normalized_matrix = [
        [pairwise_matrix[row_index][column_index] / column_sums[column_index]
         for column_index in range(criterion_count)]
        for row_index in range(criterion_count)
    ]
    criterion_weights = [mean(matrix_row) for matrix_row in normalized_matrix]
    weight_total = sum(criterion_weights)
    criterion_weights = [weight / weight_total for weight in criterion_weights]

    weighted_sum_values = [
        sum(pairwise_matrix[row_index][column_index] * criterion_weights[column_index]
            for column_index in range(criterion_count))
        for row_index in range(criterion_count)
    ]
    lambda_max = mean(
        weighted_sum_values[criterion_index] / criterion_weights[criterion_index]
        for criterion_index in range(criterion_count)
    )
    ci = (lambda_max - criterion_count) / (criterion_count - 1)
    cr = ci / RI
    return {"weights": criterion_weights, "lambda_max": lambda_max, "ci": ci, "cr": cr}


def get_adjacent_saaty_value(comparison_value, direction):
    scale_index = next(
        (value_index for value_index, scale_value in enumerate(Saaty_scale)
         if math.isclose(comparison_value, scale_value, abs_tol=1e-7)),
        None,
    )
    if scale_index is None:
        raise ValueError(f"Pairwise value {comparison_value} is not on the Saaty scale.")
    adjacent_index = scale_index - 1 if direction == "down" else scale_index + 1
    return Saaty_scale[adjacent_index] if 0 <= adjacent_index < len(Saaty_scale) else None


# ============================================================================
# SITE DATA AND WLC CALCULATIONS
# ============================================================================

def read_sites():
    with Site_file.open("r", newline="", encoding="utf-8-sig") as input_file:
        reader = csv.DictReader(input_file)
        available_fields = reader.fieldnames or []
        site_rows = list(reader)

    required_fields = [
        "Site_ID", "Site", "Cluster_3", *Score_fields.values(),
        "Rusle_MTH", "ELSUS_MTH", "Material_general",
    ]
    missing_fields = [field for field in required_fields if field not in available_fields]
    if missing_fields:
        raise ValueError("Missing site-table fields: " + ", ".join(missing_fields))
    return site_rows


def get_intentional_na_reason(site_row, criterion):
    if criterion == "Soil Erosion":
        method_value = str(site_row.get("Rusle_MTH", "")).strip()
        if method_value in {"Urban excl.", "Coastal excl.", "Other excl."}:
            return f"RUSLE not applicable / excluded ({method_value})"
    elif criterion == "Landslide":
        method_value = str(site_row.get("ELSUS_MTH", "")).strip()
        if method_value == "NoData>400m":
            return "ELSUS unavailable: nearest valid cell is beyond 400 m"
    elif criterion == "Asset Vulnerability":
        material_value = str(site_row.get("Material_general", "")).strip()
        if material_value == "Unknown / Not specified":
            return (
                "Asset Vulnerability not scored: site typology is functional "
                "rather than sufficiently physical/material-specific"
            )
    return None


def check_site_scores(site_row):
    valid_scores, intentional_na, unexpected_missing_fields = {}, {}, []
    for criterion in Criteria:
        score_field = Score_fields[criterion]
        score_value = convert_to_number(site_row.get(score_field))
        if score_value is None:
            na_reason = get_intentional_na_reason(site_row, criterion)
            if na_reason:
                intentional_na[criterion] = na_reason
            else:
                unexpected_missing_fields.append(score_field)
        elif 1 <= score_value <= 9:
            valid_scores[criterion] = score_value
        else:
            raise ValueError(
                f"{site_row.get('Site_ID')} has {score_field}={score_value}, outside the 1-9 scale."
            )
    return valid_scores, intentional_na, unexpected_missing_fields


def calculate_weight_sum(site_scores, criterion_weights):
    return sum(
        criterion_weights[criterion_index]
        for criterion_index, criterion in enumerate(Criteria)
        if criterion in site_scores
    )


def calculate_wlc(site_scores, criterion_weights):
    available_weight_sum = calculate_weight_sum(site_scores, criterion_weights)
    if available_weight_sum == 0:
        return None
    weighted_score_sum = sum(
        criterion_weights[criterion_index] * site_scores[criterion]
        for criterion_index, criterion in enumerate(Criteria)
        if criterion in site_scores
    )
    return weighted_score_sum / available_weight_sum


# ============================================================================
# SENSITIVITY ANALYSIS
# ============================================================================

# Chen et al. (2013) Overall Change Rate, using site counts instead of cell counts.
def calculate_ocr(baseline_classes, alternative_classes):
    baseline_counts = Counter(baseline_classes)
    alternative_counts = Counter(alternative_classes)
    ocr_value = 0.0
    undefined_classes = []

    for risk_class in Risk_classes:
        baseline_count = baseline_counts[risk_class]
        alternative_count = alternative_counts[risk_class]
        if baseline_count == 0:
            if alternative_count > 0:
                undefined_classes.append(risk_class)
        else:
            ocr_value += abs(alternative_count - baseline_count) / baseline_count

    if undefined_classes:
        ocr_note = (
            "OCR undefined under Chen Eq. 4 because baseline count is 0 "
            "but alternative count is >0 for: " + ", ".join(undefined_classes)
        )
        return None, ocr_note, baseline_counts, alternative_counts
    return ocr_value, "", baseline_counts, alternative_counts


# ============================================================================
# RUN ANALYSIS
# ============================================================================

def main():
    missing_files = [
        str(file_path) for file_path in list(AHP_files.values()) + [Site_file]
        if not file_path.exists()
    ]
    if missing_files:
        raise FileNotFoundError("Missing input files:\n" + "\n".join(missing_files))

    Output_folder.mkdir(parents=True, exist_ok=True)
    site_rows = read_sites()

    # Check missing values and identify intentional exclusions.
    data_quality_rows, checked_sites = [], []
    unexpected_missing_count = 0
    for site_row in site_rows:
        valid_scores, intentional_na, unexpected_missing_fields = check_site_scores(site_row)
        if intentional_na or unexpected_missing_fields:
            data_quality_row = {
                "Site_ID": site_row.get("Site_ID"),
                "Site": site_row.get("Site"),
                "Cluster": site_row.get("Cluster_3"),
                "Status": "UNEXPECTED MISSING" if unexpected_missing_fields else "Intentional N/A - renormalized",
                "Intentional_NA_Criteria": "; ".join(intentional_na),
                "Unexpected_Missing_Fields": "; ".join(unexpected_missing_fields),
                "Reason": " | ".join(
                    f"{criterion}: {reason}" for criterion, reason in intentional_na.items()
                ),
                "Rusle_MTH": site_row.get("Rusle_MTH"),
                "ELSUS_MTH": site_row.get("ELSUS_MTH"),
                "Material_general": site_row.get("Material_general"),
            }
            for score_field in Score_fields.values():
                data_quality_row[score_field] = site_row.get(score_field)
            data_quality_rows.append(data_quality_row)
            unexpected_missing_count += bool(unexpected_missing_fields)

        checked_sites.append({
            "source_row": site_row,
            "scores": valid_scores,
            "intentional_na": intentional_na,
            "cluster": str(site_row.get("Cluster_3", "")).strip(),
        })

    data_quality_fields = [
        "Site_ID", "Site", "Cluster", "Status", "Intentional_NA_Criteria",
        "Unexpected_Missing_Fields", "Reason", "Rusle_MTH", "ELSUS_MTH",
        "Material_general", *Score_fields.values(),
    ]
    write_csv(Output_folder / "00_Data_Quality.csv", data_quality_rows, data_quality_fields)
    if unexpected_missing_count:
        raise ValueError(
            f"{unexpected_missing_count} site(s) contain unexplained missing values. See 00_Data_Quality.csv."
        )

    # Calculate the baseline weights for each cluster.
    cluster_matrices, baseline_results, baseline_weight_rows = {}, {}, []
    for cluster, ahp_file in AHP_files.items():
        pairwise_matrix = read_ahp_matrix(ahp_file)
        ahp_results = calculate_ahp(pairwise_matrix)
        if ahp_results["cr"] >= CR_threshold:
            raise ValueError(f"Cluster {cluster} baseline CR is >= {CR_threshold}.")
        cluster_matrices[cluster] = pairwise_matrix
        baseline_results[cluster] = ahp_results

        for criterion_index, criterion in enumerate(Criteria):
            baseline_weight_rows.append({
                "Cluster": cluster,
                "Criterion": criterion,
                "Weight": ahp_results["weights"][criterion_index],
                "Lambda_Max": ahp_results["lambda_max"],
                "CI": ahp_results["ci"],
                "RI": RI,
                "CR": ahp_results["cr"],
                "CR_Accepted": True,
            })

    baseline_fields = [
        "Cluster", "Criterion", "Weight", "Lambda_Max", "CI", "RI", "CR", "CR_Accepted",
    ]
    write_csv(Output_folder / "01_Baseline_Weights.csv", baseline_weight_rows, baseline_fields)

    # Calculate baseline WLC scores.
    sites_by_cluster = defaultdict(list)
    for checked_site in checked_sites:
        try:
            cluster = int(checked_site["cluster"])
        except ValueError:
            continue
        if cluster not in baseline_results:
            continue

        cluster_weights = baseline_results[cluster]["weights"]
        baseline_wlc_score = calculate_wlc(checked_site["scores"], cluster_weights)
        if baseline_wlc_score is None:
            continue
        checked_site["baseline_wlc"] = baseline_wlc_score
        checked_site["baseline_class"] = classify_risk(baseline_wlc_score)
        checked_site["excluded_criteria"] = "; ".join(checked_site["intentional_na"])
        checked_site["baseline_weight_sum"] = calculate_weight_sum(checked_site["scores"], cluster_weights)
        sites_by_cluster[cluster].append(checked_site)

    # OAT pairwise sensitivity runs.
    sensitivity_run_rows, site_sensitivity_rows = [], []
    for cluster in sorted(cluster_matrices):
        baseline_matrix = cluster_matrices[cluster]
        cluster_baseline_results = baseline_results[cluster]
        cluster_sites = sites_by_cluster[cluster]
        baseline_classes = [checked_site["baseline_class"] for checked_site in cluster_sites]
        run_number = 0

        for row_index in range(len(Criteria)):
            for column_index in range(row_index + 1, len(Criteria)):
                baseline_comparison_value = baseline_matrix[row_index][column_index]
                for direction in ["down", "up"]:
                    perturbed_value = get_adjacent_saaty_value(baseline_comparison_value, direction)
                    if perturbed_value is None:
                        continue

                    run_number += 1
                    run_id = f"C{cluster}_R{run_number:02d}"
                    perturbed_matrix = [matrix_row[:] for matrix_row in baseline_matrix]
                    perturbed_matrix[row_index][column_index] = perturbed_value
                    perturbed_matrix[column_index][row_index] = 1 / perturbed_value
                    alternative_results = calculate_ahp(perturbed_matrix)
                    cr_accepted = alternative_results["cr"] < CR_threshold

                    sensitivity_run = {
                        "Cluster": cluster,
                        "Run_ID": run_id,
                        "Criterion_Row": Criteria[row_index],
                        "Criterion_Column": Criteria[column_index],
                        "Direction": direction,
                        "Base_IOI": format_saaty_value(baseline_comparison_value),
                        "Perturbed_IOI": format_saaty_value(perturbed_value),
                        "Base_CR": cluster_baseline_results["cr"],
                        "Alternative_CR": alternative_results["cr"],
                        "CR_Accepted": cr_accepted,
                    }
                    for criterion_index, criterion in enumerate(Criteria):
                        sensitivity_run[f"Weight_{criterion.replace(' ', '_')}"] = (
                            alternative_results["weights"][criterion_index]
                        )

                    if not cr_accepted:
                        sensitivity_run.update({
                            "Sites_Analyzed": 0,
                            "Sites_Changing_Class": None,
                            "Site_Class_Change_Pct": None,
                            "OCR_Chen": None,
                            "OCR_Note": "CR rejected; WLC sensitivity not evaluated.",
                        })
                        sensitivity_run_rows.append(sensitivity_run)
                        continue

                    alternative_classes = []
                    for checked_site in cluster_sites:
                        alternative_wlc_score = calculate_wlc(
                            checked_site["scores"], alternative_results["weights"]
                        )
                        alternative_risk_class = classify_risk(alternative_wlc_score)
                        alternative_classes.append(alternative_risk_class)
                        source_row = checked_site["source_row"]
                        site_sensitivity_rows.append({
                            "Cluster": cluster,
                            "Run_ID": run_id,
                            "Criterion_Row": Criteria[row_index],
                            "Criterion_Column": Criteria[column_index],
                            "Direction": direction,
                            "Base_IOI": format_saaty_value(baseline_comparison_value),
                            "Perturbed_IOI": format_saaty_value(perturbed_value),
                            "Site_ID": source_row.get("Site_ID"),
                            "Site": source_row.get("Site"),
                            "Excluded_Criteria": checked_site["excluded_criteria"],
                            "Baseline_Applicable_Weight_Sum": checked_site["baseline_weight_sum"],
                            "Alternative_Applicable_Weight_Sum": calculate_weight_sum(
                                checked_site["scores"], alternative_results["weights"]
                            ),
                            "Baseline_WLC": checked_site["baseline_wlc"],
                            "Alternative_WLC": alternative_wlc_score,
                            "Baseline_Class": checked_site["baseline_class"],
                            "Alternative_Class": alternative_risk_class,
                            "Class_Changed": checked_site["baseline_class"] != alternative_risk_class,
                        })

                    sites_changing_class = sum(
                        baseline_class != alternative_class
                        for baseline_class, alternative_class in zip(baseline_classes, alternative_classes)
                    )
                    ocr_value, ocr_note, baseline_counts, alternative_counts = calculate_ocr(
                        baseline_classes, alternative_classes
                    )
                    sensitivity_run.update({
                        "Sites_Analyzed": len(cluster_sites),
                        "Sites_Changing_Class": sites_changing_class,
                        "Site_Class_Change_Pct": 100 * sites_changing_class / len(cluster_sites),
                        "OCR_Chen": ocr_value,
                        "OCR_Note": ocr_note,
                    })
                    for risk_class in Risk_classes:
                        class_field_name = risk_class.replace(" ", "_")
                        sensitivity_run[f"Base_{class_field_name}"] = baseline_counts[risk_class]
                        sensitivity_run[f"Alt_{class_field_name}"] = alternative_counts[risk_class]
                    sensitivity_run_rows.append(sensitivity_run)

    sensitivity_run_fields = [
        "Cluster", "Run_ID", "Criterion_Row", "Criterion_Column", "Direction",
        "Base_IOI", "Perturbed_IOI", "Base_CR", "Alternative_CR", "CR_Accepted",
        *[f"Weight_{criterion.replace(' ', '_')}" for criterion in Criteria],
        "Sites_Analyzed", "Sites_Changing_Class", "Site_Class_Change_Pct", "OCR_Chen", "OCR_Note",
        *[f"Base_{risk_class.replace(' ', '_')}" for risk_class in Risk_classes],
        *[f"Alt_{risk_class.replace(' ', '_')}" for risk_class in Risk_classes],
    ]
    write_csv(Output_folder / "02_Pairwise_Sensitivity_Runs.csv", sensitivity_run_rows, sensitivity_run_fields)

    site_result_fields = [
        "Cluster", "Run_ID", "Criterion_Row", "Criterion_Column", "Direction",
        "Base_IOI", "Perturbed_IOI", "Site_ID", "Site", "Excluded_Criteria",
        "Baseline_Applicable_Weight_Sum", "Alternative_Applicable_Weight_Sum",
        "Baseline_WLC", "Alternative_WLC", "Baseline_Class", "Alternative_Class", "Class_Changed",
    ]
    write_csv(Output_folder / "03_Site_Level_Sensitivity.csv", site_sensitivity_rows, site_result_fields)

    # Criterion-level sensitivity summary.
    criterion_summary_rows = []
    for cluster in sorted(cluster_matrices):
        accepted_runs = [
            sensitivity_run for sensitivity_run in sensitivity_run_rows
            if sensitivity_run["Cluster"] == cluster and sensitivity_run["CR_Accepted"]
        ]
        for criterion in Criteria:
            criterion_runs = [
                sensitivity_run for sensitivity_run in accepted_runs
                if criterion in {sensitivity_run["Criterion_Row"], sensitivity_run["Criterion_Column"]}
            ]
            ocr_values = [
                sensitivity_run["OCR_Chen"] for sensitivity_run in criterion_runs
                if sensitivity_run["OCR_Chen"] is not None
            ]
            class_change_percentages = [
                sensitivity_run["Site_Class_Change_Pct"] for sensitivity_run in criterion_runs
                if sensitivity_run["Site_Class_Change_Pct"] is not None
            ]
            criterion_summary_rows.append({
                "Cluster": cluster,
                "Criterion": criterion,
                "Valid_Runs": len(criterion_runs),
                "Mean_OCR_Chen": mean(ocr_values) if ocr_values else None,
                "Max_OCR_Chen": max(ocr_values) if ocr_values else None,
                "Mean_Class_Change_Pct": mean(class_change_percentages) if class_change_percentages else None,
                "Max_Class_Change_Pct": max(class_change_percentages) if class_change_percentages else None,
            })

    criterion_summary_fields = [
        "Cluster", "Criterion", "Valid_Runs", "Mean_OCR_Chen", "Max_OCR_Chen",
        "Mean_Class_Change_Pct", "Max_Class_Change_Pct",
    ]
    write_csv(
        Output_folder / "04_Criterion_Sensitivity_Summary.csv",
        criterion_summary_rows,
        criterion_summary_fields,
    )

    # Short run record.
    with (Output_folder / "Run_Log.txt").open("w", encoding="utf-8") as log_file:
        log_file.write("THRACE AHP SENSITIVITY ANALYSIS\n")
        log_file.write("==================================\n\n")
        log_file.write(f"Run: {datetime.now().isoformat(timespec='seconds')}\n")
        log_file.write(f"Sites in source CSV: {len(site_rows)}\n")
        log_file.write(f"Sites with documented intentional N/A values: {len(data_quality_rows)}\n")
        log_file.write("Sites with unexpected missing criterion values: 0\n")
        log_file.write("Perturbation direction: both; adjacent steps: 1\n\n")
        for cluster in sorted(cluster_matrices):
            log_file.write(
                f"Cluster {cluster}: baseline CR={baseline_results[cluster]['cr']:.6f}; "
                f"eligible sites={len(sites_by_cluster[cluster])}\n"
            )
        accepted_run_count = sum(
            bool(sensitivity_run["CR_Accepted"]) for sensitivity_run in sensitivity_run_rows
        )
        log_file.write(f"\nSensitivity runs: {len(sensitivity_run_rows)}\n")
        log_file.write(f"CR accepted: {accepted_run_count}\n")
        log_file.write(f"CR rejected: {len(sensitivity_run_rows) - accepted_run_count}\n")

    print("Sensitivity analysis complete.")
    print(f"Sites analyzed: {len(site_rows)}")
    print(f"Sensitivity runs: {len(sensitivity_run_rows)}")
    print(f"Results saved to: {Output_folder}")


if __name__ == "__main__":
    main()
