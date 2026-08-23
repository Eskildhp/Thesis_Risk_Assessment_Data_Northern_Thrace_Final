# Monte Carlo uncertainty analysis for the Northern Thrace risk model.
# Tests uncertainty in the cluster-specific AHP weights using 1,000 simulations.

import csv
import math
import random
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean, pstdev

from openpyxl import load_workbook


# ============================================================================
# FILE PATHS AND VARIABLES
# ============================================================================

Script_folder = Path(__file__).resolve().parent
Analysis_folder = Script_folder.parent
Repository_folder = Analysis_folder.parent

AHP_files = {
    1: Repository_folder / "AHP" / "matrices" / "Thrace_AHP_C1.xlsx",
    2: Repository_folder / "AHP" / "matrices" / "Thrace_AHP_C2.xlsx",
    3: Repository_folder / "AHP" / "matrices" / "Thrace_AHP_C3.xlsx",
}

Site_file = Analysis_folder / "inputs" / "Sites_uncertainty_input.csv"
Output_folder = Analysis_folder / "results"

Criteria = [
    "Seismic", "Wildfire", "Flood", "Soil Erosion",
    "Landslide", "Asset Vulnerability",
]

Score_fields = {
    "Seismic": "SEIS_SCR",
    "Wildfire": "FWI_SCR",
    "Flood": "FLOOD_SCR",
    "Soil Erosion": "RUSLE_SCR",
    "Landslide": "ELSUS_SCR",
    "Asset Vulnerability": "ASSET_SCR",
}

Baseline_WLC_field = "WLC_RISK2"
Iterations = 1000
Perturbation_fraction = 0.10
Random_seed = 20260813
Convergence_checkpoints = [100, 250, 500, 750, 1000]
Baseline_validation_tolerance = 1e-5
RI = 1.24
Risk_classes = ["Very Low", "Low", "Moderate", "High", "Very High"]


# ============================================================================
# GENERAL FUNCTIONS
# ============================================================================

def convert_to_number(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def classify_risk(score):
    if score < 1 or score > 9:
        raise ValueError(f"WLC score {score} is outside the expected 1-9 range.")
    if score < 2:
        return "Very Low"
    if score < 4:
        return "Low"
    if score < 6:
        return "Moderate"
    if score < 8:
        return "High"
    return "Very High"


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


# ============================================================================
# AHP MATRIX AND WEIGHT CALCULATIONS
# ============================================================================

def read_ahp_matrix(ahp_file):
    workbook = load_workbook(ahp_file, data_only=True)
    sheet = workbook.active
    header_row = start_column = None

    for row_number in range(1, min(sheet.max_row, 25) + 1):
        for column_number in range(1, sheet.max_column - len(Criteria) + 2):
            headings = [
                sheet.cell(row_number, column_number + criterion_index).value
                for criterion_index in range(len(Criteria))
            ]
            if headings == Criteria:
                header_row, start_column = row_number, column_number
                break
        if header_row is not None:
            break

    if header_row is None or start_column is None:
        raise ValueError(f"Could not find the AHP criterion headings in {ahp_file.name}.")

    row_labels = [
        sheet.cell(header_row + 1 + criterion_index, start_column - 1).value
        for criterion_index in range(len(Criteria))
    ]
    if row_labels != Criteria:
        raise ValueError(f"Criterion rows do not match the expected order in {ahp_file.name}.")

    pairwise_matrix = []
    for row_index in range(len(Criteria)):
        matrix_row = [
            convert_to_number(sheet.cell(
                header_row + 1 + row_index,
                start_column + column_index
            ).value)
            for column_index in range(len(Criteria))
        ]
        if None in matrix_row:
            raise ValueError(f"Blank matrix value found in {ahp_file.name}.")
        pairwise_matrix.append(matrix_row)

    for row_index in range(len(Criteria)):
        if not math.isclose(pairwise_matrix[row_index][row_index], 1.0, abs_tol=1e-8):
            raise ValueError(f"The diagonal of {ahp_file.name} contains a value other than 1.")
        for column_index in range(row_index + 1, len(Criteria)):
            reciprocal = (
                pairwise_matrix[row_index][column_index]
                * pairwise_matrix[column_index][row_index]
            )
            if not math.isclose(reciprocal, 1.0, abs_tol=1e-7):
                raise ValueError(f"The AHP matrix in {ahp_file.name} is not reciprocal.")

    return pairwise_matrix


def calculate_ahp(pairwise_matrix):
    criterion_count = len(pairwise_matrix)
    column_sums = [
        sum(pairwise_matrix[row_index][column_index] for row_index in range(criterion_count))
        for column_index in range(criterion_count)
    ]
    normalized_matrix = [
        [
            pairwise_matrix[row_index][column_index] / column_sums[column_index]
            for column_index in range(criterion_count)
        ]
        for row_index in range(criterion_count)
    ]

    weights = [mean(row) for row in normalized_matrix]
    weight_total = sum(weights)
    weights = [weight / weight_total for weight in weights]

    weighted_sums = [
        sum(
            pairwise_matrix[row_index][column_index] * weights[column_index]
            for column_index in range(criterion_count)
        )
        for row_index in range(criterion_count)
    ]
    consistency_values = [
        weighted_sums[criterion_index] / weights[criterion_index]
        for criterion_index in range(criterion_count)
    ]

    lambda_max = mean(consistency_values)
    ci = (lambda_max - criterion_count) / (criterion_count - 1)
    cr = ci / RI

    return {"weights": weights, "lambda_max": lambda_max, "ci": ci, "cr": cr}


# ============================================================================
# SITE DATA AND WLC
# ============================================================================

def read_sites():
    with Site_file.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        fields = reader.fieldnames or []
        sites = list(reader)

    required_fields = [
        "Site_ID", "Site", "Cluster_3", *Score_fields.values(),
        "Rusle_MTH", "ELSUS_MTH", "Material_general", Baseline_WLC_field,
    ]
    missing_fields = [field for field in required_fields if field not in fields]
    if missing_fields:
        raise ValueError("Missing site-table fields: " + ", ".join(missing_fields))
    return sites


def intentional_na_reason(site, criterion):
    if criterion == "Soil Erosion":
        method = str(site.get("Rusle_MTH", "")).strip()
        if method in {"Urban excl.", "Coastal excl.", "Other excl."}:
            return f"RUSLE excluded ({method})"

    if criterion == "Landslide":
        if str(site.get("ELSUS_MTH", "")).strip() == "NoData>400m":
            return "ELSUS unavailable: nearest valid cell beyond 400 m"

    if criterion == "Asset Vulnerability":
        if str(site.get("Material_general", "")).strip() == "Unknown / Not specified":
            return "Asset Vulnerability not scored: insufficient physical/material information"

    return None


def check_site_scores(site):
    scores, intentional_na, unexpected_missing_fields = {}, {}, []

    for criterion in Criteria:
        field = Score_fields[criterion]
        value = convert_to_number(site.get(field))

        if value is None:
            reason = intentional_na_reason(site, criterion)
            if reason is not None:
                intentional_na[criterion] = reason
            else:
                unexpected_missing_fields.append(field)
        elif 1 <= value <= 9:
            scores[criterion] = value
        else:
            raise ValueError(
                f"{site.get('Site_ID')} has {field}={value}, outside the expected 1-9 range."
            )

    return scores, intentional_na, unexpected_missing_fields


def calculate_weight_sum(scores, weights):
    return sum(
        weights[criterion_index]
        for criterion_index, criterion in enumerate(Criteria)
        if criterion in scores
    )


def calculate_wlc(scores, weights):
    available_weight_sum = calculate_weight_sum(scores, weights)
    if available_weight_sum <= 0:
        return None

    weighted_score_sum = sum(
        weights[criterion_index] * scores[criterion]
        for criterion_index, criterion in enumerate(Criteria)
        if criterion in scores
    )
    return weighted_score_sum / available_weight_sum


# ============================================================================
# MONTE CARLO SIMULATION
# ============================================================================

def perturb_weights(baseline_weights, random_generator):
    perturbed_weights = [
        weight * random_generator.uniform(
            1.0 - Perturbation_fraction,
            1.0 + Perturbation_fraction,
        )
        for weight in baseline_weights
    ]
    total_weight = sum(perturbed_weights)
    return [weight / total_weight for weight in perturbed_weights]


def calculate_simulation_statistics(values):
    mean_value = mean(values)
    standard_deviation = pstdev(values)
    return {
        "mean": mean_value,
        "sd": standard_deviation,
        "cv_pct": 100.0 * standard_deviation / mean_value if mean_value else None,
        "min": min(values),
        "max": max(values),
    }


# ============================================================================
# RUN ANALYSIS
# ============================================================================

def main():
    missing_files = [
        str(path) for path in list(AHP_files.values()) + [Site_file]
        if not path.exists()
    ]
    if missing_files:
        raise FileNotFoundError("Missing input files:\n" + "\n".join(missing_files))

    Output_folder.mkdir(parents=True, exist_ok=True)
    sites = read_sites()

    # Check missing values and identify intentional exclusions.
    data_quality_rows = []
    checked_sites = []
    unexpected_missing_count = 0

    for site in sites:
        scores, intentional_na, unexpected_missing_fields = check_site_scores(site)

        if intentional_na or unexpected_missing_fields:
            data_quality_row = {
                "Site_ID": site.get("Site_ID"),
                "Site": site.get("Site"),
                "Cluster": site.get("Cluster_3"),
                "Status": (
                    "UNEXPECTED MISSING"
                    if unexpected_missing_fields
                    else "Intentional N/A - renormalized"
                ),
                "Intentional_NA_Criteria": "; ".join(intentional_na),
                "Unexpected_Missing_Fields": "; ".join(unexpected_missing_fields),
                "Reason": " | ".join(
                    f"{criterion}: {reason}"
                    for criterion, reason in intentional_na.items()
                ),
                "Rusle_MTH": site.get("Rusle_MTH"),
                "ELSUS_MTH": site.get("ELSUS_MTH"),
                "Material_general": site.get("Material_general"),
            }
            for field in Score_fields.values():
                data_quality_row[field] = site.get(field)
            data_quality_rows.append(data_quality_row)

            if unexpected_missing_fields:
                unexpected_missing_count += 1

        checked_sites.append({
            "raw": site,
            "scores": scores,
            "intentional_na": intentional_na,
        })

    data_quality_fields = [
        "Site_ID", "Site", "Cluster", "Status", "Intentional_NA_Criteria",
        "Unexpected_Missing_Fields", "Reason", "Rusle_MTH", "ELSUS_MTH",
        "Material_general", *Score_fields.values(),
    ]
    write_csv(Output_folder / "00_Data_Quality.csv", data_quality_rows, data_quality_fields)

    if unexpected_missing_count:
        raise ValueError(
            f"{unexpected_missing_count} site(s) contain unexplained missing values. "
            "See 00_Data_Quality.csv."
        )

    # Calculate the baseline AHP weights for each cluster.
    baseline_results = {}
    baseline_weight_rows = []

    for cluster, ahp_file in AHP_files.items():
        pairwise_matrix = read_ahp_matrix(ahp_file)
        ahp_results = calculate_ahp(pairwise_matrix)
        baseline_results[cluster] = ahp_results

        for criterion_index, criterion in enumerate(Criteria):
            baseline_weight_rows.append({
                "Cluster": cluster,
                "Criterion": criterion,
                "Weight": ahp_results["weights"][criterion_index],
                "Lambda_Max": ahp_results["lambda_max"],
                "CI": ahp_results["ci"],
                "RI": RI,
                "CR": ahp_results["cr"],
            })

    baseline_weight_fields = [
        "Cluster", "Criterion", "Weight", "Lambda_Max", "CI", "RI", "CR"
    ]
    write_csv(
        Output_folder / "01_Baseline_Weights.csv",
        baseline_weight_rows,
        baseline_weight_fields,
    )

    # Calculate baseline WLC scores.
    sites_by_cluster = defaultdict(list)

    for checked_site in checked_sites:
        raw_site = checked_site["raw"]
        try:
            cluster = int(str(raw_site.get("Cluster_3", "")).strip())
        except (TypeError, ValueError):
            continue

        if cluster not in baseline_results:
            continue

        baseline_weights = baseline_results[cluster]["weights"]
        baseline_wlc = calculate_wlc(checked_site["scores"], baseline_weights)
        if baseline_wlc is None:
            continue

        existing_wlc = convert_to_number(raw_site.get(Baseline_WLC_field))

        site_record = dict(checked_site)
        site_record.update({
            "cluster": cluster,
            "baseline_wlc": baseline_wlc,
            "baseline_class": classify_risk(baseline_wlc),
            "existing_wlc": existing_wlc,
            "existing_delta": None if existing_wlc is None else baseline_wlc - existing_wlc,
            "excluded_criteria": "; ".join(checked_site["intentional_na"]),
            "baseline_weight_sum": calculate_weight_sum(
                checked_site["scores"], baseline_weights
            ),
        })
        sites_by_cluster[cluster].append(site_record)

    # Validate the calculated baseline against the ArcGIS WLC result.
    baseline_validation_rows = []
    baseline_mismatch_count = 0

    for cluster in sorted(sites_by_cluster):
        for site_record in sites_by_cluster[cluster]:
            raw_site = site_record["raw"]
            existing_wlc = site_record["existing_wlc"]
            calculated_wlc = site_record["baseline_wlc"]

            if existing_wlc is None:
                difference = absolute_difference = None
                passed = False
                reason = f"{Baseline_WLC_field} is blank"
            else:
                difference = calculated_wlc - existing_wlc
                absolute_difference = abs(difference)
                passed = absolute_difference <= Baseline_validation_tolerance
                reason = "" if passed else "Calculated baseline differs from ArcGIS baseline"

            validation_row = {
                "Cluster": cluster,
                "Site_ID": raw_site.get("Site_ID"),
                "Site": raw_site.get("Site"),
                "Calculated_Baseline_WLC": calculated_wlc,
                "ArcGIS_Baseline_WLC": existing_wlc,
                "Delta": difference,
                "Abs_Delta": absolute_difference,
                "Tolerance": Baseline_validation_tolerance,
                "Passed": passed,
                "Reason": reason,
            }
            baseline_validation_rows.append(validation_row)
            if not passed:
                baseline_mismatch_count += 1

    baseline_validation_fields = [
        "Cluster", "Site_ID", "Site", "Calculated_Baseline_WLC",
        "ArcGIS_Baseline_WLC", "Delta", "Abs_Delta", "Tolerance", "Passed", "Reason",
    ]
    write_csv(
        Output_folder / "01A_Baseline_Validation.csv",
        baseline_validation_rows,
        baseline_validation_fields,
    )

    if baseline_mismatch_count:
        raise ValueError(
            f"{baseline_mismatch_count} of {len(baseline_validation_rows)} sites "
            f"do not match {Baseline_WLC_field} within ±{Baseline_validation_tolerance}. "
            "See 01A_Baseline_Validation.csv."
        )

    # Prepare storage for simulation results.
    random_generator = random.Random(Random_seed)
    iteration_weight_rows = []
    simulated_wlc_values = {}
    simulated_class_counts = {}
    convergence_rows = []

    for cluster, cluster_sites in sites_by_cluster.items():
        for site_record in cluster_sites:
            site_key = (cluster, str(site_record["raw"].get("Site_ID")))
            simulated_wlc_values[site_key] = []
            simulated_class_counts[site_key] = Counter()

    # Run the simulations separately for each cluster.
    for cluster in sorted(sites_by_cluster):
        cluster_sites = sites_by_cluster[cluster]
        baseline_weights = baseline_results[cluster]["weights"]
        checkpoint_results = {}
        previous_checkpoint = None

        for iteration in range(1, Iterations + 1):
            alternative_weights = perturb_weights(baseline_weights, random_generator)

            iteration_weight_row = {"Cluster": cluster, "Iteration": iteration}
            for criterion_index, criterion in enumerate(Criteria):
                iteration_weight_row[
                    f"Weight_{criterion.replace(' ', '_')}"
                ] = alternative_weights[criterion_index]
            iteration_weight_rows.append(iteration_weight_row)

            for site_record in cluster_sites:
                site_key = (cluster, str(site_record["raw"].get("Site_ID")))
                simulated_wlc = calculate_wlc(site_record["scores"], alternative_weights)
                simulated_class = classify_risk(simulated_wlc)
                simulated_wlc_values[site_key].append(simulated_wlc)
                simulated_class_counts[site_key][simulated_class] += 1

            if iteration in Convergence_checkpoints:
                current_means = {}
                current_standard_deviations = {}

                for site_record in cluster_sites:
                    site_key = (cluster, str(site_record["raw"].get("Site_ID")))
                    values = simulated_wlc_values[site_key]
                    current_means[site_key] = mean(values)
                    current_standard_deviations[site_key] = pstdev(values)

                if previous_checkpoint is None:
                    convergence_rows.append({
                        "Cluster": cluster,
                        "Checkpoint_Iterations": iteration,
                        "Compared_With": None,
                        "Mean_Abs_Change_in_Site_Mean": None,
                        "Max_Abs_Change_in_Site_Mean": None,
                        "Mean_Abs_Change_in_Site_SD": None,
                        "Max_Abs_Change_in_Site_SD": None,
                    })
                else:
                    previous_means, previous_standard_deviations = checkpoint_results[
                        previous_checkpoint
                    ]
                    mean_differences = [
                        abs(current_means[key] - previous_means[key])
                        for key in current_means
                    ]
                    sd_differences = [
                        abs(current_standard_deviations[key] - previous_standard_deviations[key])
                        for key in current_standard_deviations
                    ]
                    convergence_rows.append({
                        "Cluster": cluster,
                        "Checkpoint_Iterations": iteration,
                        "Compared_With": previous_checkpoint,
                        "Mean_Abs_Change_in_Site_Mean": mean(mean_differences),
                        "Max_Abs_Change_in_Site_Mean": max(mean_differences),
                        "Mean_Abs_Change_in_Site_SD": mean(sd_differences),
                        "Max_Abs_Change_in_Site_SD": max(sd_differences),
                    })

                checkpoint_results[iteration] = (
                    dict(current_means),
                    dict(current_standard_deviations),
                )
                previous_checkpoint = iteration

    # Summarize simulation results for each site.
    site_summary_rows = []

    for cluster in sorted(sites_by_cluster):
        for site_record in sites_by_cluster[cluster]:
            raw_site = site_record["raw"]
            site_key = (cluster, str(raw_site.get("Site_ID")))
            values = simulated_wlc_values[site_key]
            statistics = calculate_simulation_statistics(values)
            class_counts = simulated_class_counts[site_key]

            baseline_class = site_record["baseline_class"]
            stability_count = class_counts[baseline_class]
            stability_percentage = 100.0 * stability_count / len(values)
            mean_risk_class = classify_risk(statistics["mean"])

            site_summary_row = {
                "Cluster": cluster,
                "Site_ID": raw_site.get("Site_ID"),
                "Site": raw_site.get("Site"),
                "Excluded_Criteria": site_record["excluded_criteria"],
                "Baseline_Applicable_Weight_Sum": site_record["baseline_weight_sum"],
                "Baseline_WLC": site_record["baseline_wlc"],
                "Existing_WLC": site_record["existing_wlc"],
                "Baseline_Check_Delta": site_record["existing_delta"],
                "Baseline_Class": baseline_class,
                "MC_Mean_WLC": statistics["mean"],
                "MC_SD": statistics["sd"],
                "MC_CV_Pct": statistics["cv_pct"],
                "MC_Min_WLC": statistics["min"],
                "MC_Max_WLC": statistics["max"],
                "MC_Mean_Delta_From_Baseline": statistics["mean"] - site_record["baseline_wlc"],
                "MC_Abs_Mean_Delta_From_Baseline": abs(
                    statistics["mean"] - site_record["baseline_wlc"]
                ),
                "MC_Mean_Class": mean_risk_class,
                "MC_Mean_Class_Changed": mean_risk_class != baseline_class,
                "Baseline_Class_Stability_Count": stability_count,
                "Baseline_Class_Stability_Pct": stability_percentage,
            }

            for risk_class in Risk_classes:
                field_name = risk_class.replace(" ", "_")
                site_summary_row[f"Count_{field_name}"] = class_counts[risk_class]
                site_summary_row[f"Pct_{field_name}"] = (
                    100.0 * class_counts[risk_class] / len(values)
                )

            site_summary_rows.append(site_summary_row)

    site_summary_fields = [
        "Cluster", "Site_ID", "Site", "Excluded_Criteria",
        "Baseline_Applicable_Weight_Sum", "Baseline_WLC", "Existing_WLC",
        "Baseline_Check_Delta", "Baseline_Class", "MC_Mean_WLC", "MC_SD",
        "MC_CV_Pct", "MC_Min_WLC", "MC_Max_WLC", "MC_Mean_Delta_From_Baseline",
        "MC_Abs_Mean_Delta_From_Baseline", "MC_Mean_Class",
        "MC_Mean_Class_Changed", "Baseline_Class_Stability_Count",
        "Baseline_Class_Stability_Pct",
        *[
            field
            for risk_class in Risk_classes
            for field in (
                f"Count_{risk_class.replace(' ', '_')}",
                f"Pct_{risk_class.replace(' ', '_')}",
            )
        ],
    ]
    write_csv(Output_folder / "03_MC_Site_Summary.csv", site_summary_rows, site_summary_fields)

    # Summarize simulation results for each cluster.
    cluster_summary_rows = []

    for cluster in sorted(sites_by_cluster):
        cluster_rows = [row for row in site_summary_rows if row["Cluster"] == cluster]
        cluster_summary_rows.append({
            "Cluster": cluster,
            "Sites_n": len(cluster_rows),
            "Mean_Baseline_WLC": mean(row["Baseline_WLC"] for row in cluster_rows),
            "Mean_MC_Mean_WLC": mean(row["MC_Mean_WLC"] for row in cluster_rows),
            "Mean_Abs_MC_Mean_Delta": mean(
                row["MC_Abs_Mean_Delta_From_Baseline"] for row in cluster_rows
            ),
            "Max_Abs_MC_Mean_Delta": max(
                row["MC_Abs_Mean_Delta_From_Baseline"] for row in cluster_rows
            ),
            "Mean_MC_SD": mean(row["MC_SD"] for row in cluster_rows),
            "Max_MC_SD": max(row["MC_SD"] for row in cluster_rows),
            "Mean_MC_CV_Pct": mean(row["MC_CV_Pct"] for row in cluster_rows),
            "Max_MC_CV_Pct": max(row["MC_CV_Pct"] for row in cluster_rows),
            "Mean_Class_Stability_Pct": mean(
                row["Baseline_Class_Stability_Pct"] for row in cluster_rows
            ),
            "Minimum_Class_Stability_Pct": min(
                row["Baseline_Class_Stability_Pct"] for row in cluster_rows
            ),
            "Sites_MC_Mean_Class_Changed": sum(
                bool(row["MC_Mean_Class_Changed"]) for row in cluster_rows
            ),
        })

    cluster_summary_fields = [
        "Cluster", "Sites_n", "Mean_Baseline_WLC", "Mean_MC_Mean_WLC",
        "Mean_Abs_MC_Mean_Delta", "Max_Abs_MC_Mean_Delta", "Mean_MC_SD",
        "Max_MC_SD", "Mean_MC_CV_Pct", "Max_MC_CV_Pct",
        "Mean_Class_Stability_Pct", "Minimum_Class_Stability_Pct",
        "Sites_MC_Mean_Class_Changed",
    ]
    write_csv(
        Output_folder / "04_MC_Cluster_Summary.csv",
        cluster_summary_rows,
        cluster_summary_fields,
    )

    convergence_fields = [
        "Cluster", "Checkpoint_Iterations", "Compared_With",
        "Mean_Abs_Change_in_Site_Mean", "Max_Abs_Change_in_Site_Mean",
        "Mean_Abs_Change_in_Site_SD", "Max_Abs_Change_in_Site_SD",
    ]
    write_csv(Output_folder / "05_MC_Convergence.csv", convergence_rows, convergence_fields)

    iteration_weight_fields = [
        "Cluster", "Iteration",
        *[f"Weight_{criterion.replace(' ', '_')}" for criterion in Criteria],
    ]
    write_csv(
        Output_folder / "02_MC_Iteration_Weights.csv",
        iteration_weight_rows,
        iteration_weight_fields,
    )

    top_uncertain_sites = sorted(
        site_summary_rows,
        key=lambda row: row["MC_CV_Pct"],
        reverse=True,
    )[:20]
    write_csv(
        Output_folder / "06_Top_20_Uncertain_Sites.csv",
        top_uncertain_sites,
        site_summary_fields,
    )

    # Write a short run log.
    with (Output_folder / "Run_Log.txt").open("w", encoding="utf-8") as log:
        log.write("THRACE AHP MONTE CARLO UNCERTAINTY ANALYSIS\n")
        log.write("============================================\n\n")
        log.write(f"Run: {datetime.now().isoformat(timespec='seconds')}\n")
        log.write(f"Iterations per cluster: {Iterations}\n")
        log.write(
            f"Weight perturbation: ±{Perturbation_fraction * 100:.0f}% uniform multiplicative\n"
        )
        log.write(f"Random seed: {Random_seed}\n")
        log.write("Intentional N/A: excluded site-by-site; remaining weights renormalized\n")
        log.write(f"Baseline validation field: {Baseline_WLC_field}\n")
        log.write(f"Baseline validation tolerance: {Baseline_validation_tolerance}\n")
        log.write(
            f"Baseline validation mismatches: {baseline_mismatch_count} "
            f"of {len(baseline_validation_rows)}\n\n"
        )
        for cluster in sorted(sites_by_cluster):
            log.write(f"Cluster {cluster}: {len(sites_by_cluster[cluster])} sites\n")

        log.write("\nOutputs:\n")
        for output_name in [
            "00_Data_Quality.csv",
            "01_Baseline_Weights.csv",
            "01A_Baseline_Validation.csv",
            "02_MC_Iteration_Weights.csv",
            "03_MC_Site_Summary.csv",
            "04_MC_Cluster_Summary.csv",
            "05_MC_Convergence.csv",
            "06_Top_20_Uncertain_Sites.csv",
        ]:
            log.write(f"  {output_name}\n")

    print("Monte Carlo uncertainty analysis complete.")
    print(f"Sites analyzed: {len(sites)}")
    print(f"Iterations per cluster: {Iterations}")
    print(f"Results saved to: {Output_folder}")


if __name__ == "__main__":
    main()
